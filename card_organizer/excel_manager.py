# my_card_app/excel_manager.py

import os
import win32com.client as win32
from openpyxl import load_workbook, Workbook
from tkinter import messagebox # Only for critical errors that can't be handled by the caller

# Import configurations
from . import config # Relative import

def create_excel_file():
    """
    Creates a new Excel file with headers if it doesn't exist.
    Returns True on success, False on failure (e.g., permission issues).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"
    ws.append(config.ALL_FIELDS) # Use ALL_FIELDS for the header
    try:
        wb.save(config.EXCEL_FILE_PATH)
        print(f"DEBUG: Created new Excel file: {config.EXCEL_FILE_PATH}")
        return True
    except Exception as e:
        messagebox.showerror("File Creation Error",
                             f"Could not create Excel file at '{config.EXCEL_FILE_PATH}'. "
                             f"Check permissions or path.\nError: {e}")
        return False

def get_all_excel_data():
    """
    Reads all data from the Excel file (skipping the header row).
    Returns a list of lists, or an empty list if no data or file creation fails.
    """
    try:
        if not os.path.exists(config.EXCEL_FILE_PATH):
            if not create_excel_file(): # Try to create if it doesn't exist
                return [] # If creation fails, return empty
            
        wb = load_workbook(config.EXCEL_FILE_PATH)
        ws = wb.active
        data = []
        # min_row=2 skips the header row (Excel row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        return data
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Excel file:\n{e}")
        return []

def append_to_file(data_row):
    """
    Appends a new row of data to the Excel file using openpyxl.
    This method works when Excel is closed.
    Returns True on success.
    Returns False on PermissionError (file locked by another process, e.g., Excel open).
    Returns False on other Exceptions (with a messagebox shown).
    """
    try:
        if not os.path.exists(config.EXCEL_FILE_PATH):
            if not create_excel_file():
                return False # Failed to create file, cannot append

        wb = load_workbook(config.EXCEL_FILE_PATH)
        ws = wb.active

        ws.append(data_row)
        wb.save(config.EXCEL_FILE_PATH)
        print("DEBUG: Appended data using openpyxl.")
        return True
    except PermissionError:
        print("DEBUG: openpyxl failed due to PermissionError (file likely open).")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Failed to write to file (openpyxl):\n{e}")
        return False

def update_excel_row_openpyxl(row_index, new_data):
    """
    Updates a specific row in the Excel file using openpyxl.
    This method works when Excel is closed.
    `row_index` is the 0-based index of the data row in the list returned by get_all_excel_data().
    Returns True on success.
    Returns False on PermissionError (file locked).
    Returns False on other Exceptions (with a messagebox shown).
    """
    try:
        if not os.path.exists(config.EXCEL_FILE_PATH):
            messagebox.showerror("Error", "Excel file does not exist to update.")
            return False
        wb = load_workbook(config.EXCEL_FILE_PATH)
        ws = wb.active

        # openpyxl is 1-indexed for rows. We skip header (row 1), so data[0] is Excel row 2.
        # Therefore, list_index + 2 maps to the correct Excel sheet row number.
        excel_row_num = row_index + 2
        
        if excel_row_num > ws.max_row:
            messagebox.showwarning("Warning", f"Attempted to update row {row_index} which is out of bounds.")
            return False

        for col_idx, value in enumerate(new_data, start=1):
            ws.cell(row=excel_row_num, column=col_idx, value=value)

        wb.save(config.EXCEL_FILE_PATH)
        print(f"DEBUG: Updated row {excel_row_num} using openpyxl.")
        return True
    except PermissionError:
        print(f"DEBUG: openpyxl update failed due to PermissionError (file likely open).")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Failed to update row (openpyxl):\n{e}")
        return False

def find_first_empty_row_com(sheet):
    """Helper for win32com to find the first empty row in an Excel sheet."""
    row = 1
    while True:
        cell_value = sheet.Cells(row, 1).Value
        if cell_value is None or str(cell_value).strip() == "":
            return row
        row += 1

def append_to_open_excel_com(data_row):
    """
    Appends a new row to an already open Excel file using win32com.
    This method works when Excel is already open.
    Returns True on success, False on failure.
    """
    try:
        print("DEBUG: Attempting to connect to Excel via COM for appending...")
        try:
            excel = win32.GetActiveObject('Excel.Application')
            print("DEBUG: Connected to active Excel instance.")
        except Exception:
            excel = win32.Dispatch('Excel.Application')
            print("DEBUG: Dispatched new Excel instance.")
        
        excel.Visible = True # Make Excel visible for debugging purposes

        workbook = None
        # Iterate through all open workbooks to find the target file
        for wb in excel.Workbooks:
            print(f"DEBUG: Checking open workbook: {wb.FullName.lower()}")
            if wb.FullName.lower() == config.EXCEL_FILE_PATH.lower():
                workbook = wb
                break
        
        if workbook is None:
            print(f"DEBUG: Target workbook '{config.EXCEL_FILE_PATH}' not found open in Excel.")
            # If not found, try to open it
            try:
                workbook = excel.Workbooks.Open(config.EXCEL_FILE_PATH)
                print(f"DEBUG: Successfully opened workbook via COM: {config.EXCEL_FILE_PATH}")
            except Exception as e_open:
                print(f"DEBUG: Failed to open workbook via COM: {e_open}")
                return False # Cannot find or open the workbook

        print(f"DEBUG: Workbook found/opened: {workbook.FullName}")
        
        if workbook.ReadOnly:
            print(f"DEBUG: Workbook is open in Read-Only mode via COM. Cannot save changes.")
            return False # Cannot save if read-only

        sheet = workbook.Sheets(1) # Assuming data is on the first sheet (index 1 in COM)
        print(f"DEBUG: Connected to sheet: {sheet.Name}")
        next_row = find_first_empty_row_com(sheet)
        print(f"DEBUG: Next empty row for append: {next_row}")

        for col, value in enumerate(data_row, start=1):
            sheet.Cells(next_row, col).Value = value
        
        print(f"DEBUG: Data written to cells. Attempting to save workbook...")
        workbook.Save()
        print("DEBUG: Workbook saved successfully via COM.")
        return True
    except Exception as e:
        print(f"DEBUG: append_to_open_excel_com failed with exception: {e}")
        return False

def update_excel_row_com(row_index, new_data):
    """
    Updates a specific row in an already open Excel file using win32com.
    This method works when Excel is already open.
    `row_index` is the 0-based index of the data row in the list returned by get_all_excel_data().
    Returns True on success, False on failure.
    """
    try:
        print("DEBUG: Attempting to connect to Excel via COM for updating...")
        try:
            excel = win32.GetActiveObject('Excel.Application')
            print("DEBUG: Connected to active Excel instance.")
        except Exception:
            excel = win32.Dispatch('Excel.Application')
            print("DEBUG: Dispatched new Excel instance.")
        
        excel.Visible = True # Make Excel visible for debugging

        workbook = None
        for wb in excel.Workbooks:
            print(f"DEBUG: Checking open workbook: {wb.FullName.lower()}")
            if wb.FullName.lower() == config.EXCEL_FILE_PATH.lower():
                workbook = wb
                break
        
        if workbook is None:
            print(f"DEBUG: Target workbook '{config.EXCEL_FILE_PATH}' not found open in Excel.")
            # If not found, try to open it
            try:
                workbook = excel.Workbooks.Open(config.EXCEL_FILE_PATH)
                print(f"DEBUG: Successfully opened workbook via COM: {config.EXCEL_FILE_PATH}")
            except Exception as e_open:
                print(f"DEBUG: Failed to open workbook via COM: {e_open}")
                return False # Cannot find or open the workbook

        print(f"DEBUG: Workbook found/opened: {workbook.FullName}")
        
        if workbook.ReadOnly:
            print(f"DEBUG: Workbook is open in Read-Only mode via COM. Cannot save changes.")
            return False # Cannot save if read-only

        sheet = workbook.Sheets(1)
        print(f"DEBUG: Connected to sheet: {sheet.Name}")
        excel_row_num = row_index + 2 # 0-indexed list -> 1-indexed sheet row + 1 for header
        print(f"DEBUG: Updating Excel row: {excel_row_num}")

        for col_idx, value in enumerate(new_data, start=1):
            sheet.Cells(excel_row_num, col_idx).Value = value
        
        print(f"DEBUG: Data written to cells. Attempting to save workbook...")
        workbook.Save()
        print("DEBUG: Workbook saved successfully via COM.")
        return True
    except Exception as e:
        print(f"DEBUG: update_excel_row_com failed with exception: {e}")
        return False